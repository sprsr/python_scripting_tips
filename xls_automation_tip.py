#################################################################################################
# Helpful Excel Automation Library Tip
# @sprsr
#################################################################################################
# Heres a nice little tip if performing operations on excel files
# Previously, I always liked to use xlrd and xlwt,
# but recently discovered an alternative.
# Lets say we want to save some columns and delete them
# using xlrd and xlwt we need to open up the workbook twice...
import xlrd
import xlwt

# Open the Excel workbook
workbook = xlrd.open_workbook('input.xls')

# Select the sheet by index or name
sheet = workbook.sheet_by_index(0)

# Create an empty dictionary
data_dict = {}

# Iterate through each row
for row_index in range(sheet.nrows):
    column_a_value = sheet.cell_value(row_index, 0)  # Assuming column A is the first column
    column_b_value = sheet.cell_value(row_index, 1)  # Assuming column B is the second column
    data_dict[column_a_value] = column_b_value

# Create a new workbook
new_workbook = xlwt.Workbook()

# Add a new sheet to the workbook
new_sheet = new_workbook.add_sheet('Sheet1')

# Iterate through the dictionary and write entries to the new sheet
row_index = 0
for key, value in data_dict.items():
    new_sheet.write(row_index, 0, key)
    new_sheet.write(row_index, 1, value)
    row_index += 1

# Save the new workbook
new_workbook.save('output.xls')

# Now...
# If we use this library,
# openpyxl
# We can greatly improve the length of the operation
# We no longer have to open the workbook twice
# (Once to read, Once to write)
import openpyxl

# Open the Excel workbook
wb = openpyxl.load_workbook('input.xlsx')

# Select the active sheet (or specify the sheet name)
sheet = wb.active

# Create an empty dictionary
data_dict = {}

# Iterate through each row
for row in sheet.iter_rows(min_row=2, values_only=True):
    column_a_value = row[0]  # Assuming column A is the first column
    column_b_value = row[1]  # Assuming column B is the second column
    data_dict[column_a_value] = column_b_value

# Delete columns A and B
sheet.delete_cols(1, 2)

# Save the modified workbook as a new workbook
wb.save('output.xlsx')
